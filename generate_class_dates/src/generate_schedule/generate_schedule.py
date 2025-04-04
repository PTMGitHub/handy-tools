import pandas as pd
from datetime import datetime, timedelta
from openpyxl import load_workbook


def generate_schedule(year):
    start_date = datetime(year, 4, 1)  # Start from April 1st
    end_date = datetime(year + 1, 3, 31)  # End on March 31st next year

    # Define class schedule
    classes = [
        ("Monday", "18:30", 60, "Pump", "Nuffield"),
        ("Tuesday", "18:00", 45, "Pump", "JD"),
        ("Tuesday", "19:00", 45, "Combat", "JD"),
        ("Friday", "17:00", 45, "Combat", "Nuffield"),
        ("Friday", "18:00", 60, "Pump", "Nuffield"),
    ]

    # Create a list to store schedule data
    schedule = []
    current_date = start_date

    while current_date <= end_date:
        weekday = current_date.strftime('%A')
        for day, time, duration, prog, club in classes:
            if weekday == day:
                schedule.append([
                    current_date, time, duration, prog, club, "Own", "", 22.50  # Ensure Rate is a number
                ])
        current_date += timedelta(days=1)

    # Create DataFrame
    columns = ["Date", "Time", "Duration", "Prog", "Club", "Cover/Own", "Covered For", "Rate"]
    df = pd.DataFrame(schedule, columns=columns)

    # Save to Excel
    file_name = f"Exercise_Class_Schedule_{year}-{year + 1}.xlsx"
    df.to_excel(file_name, index=False, engine='openpyxl')

    # Load workbook to format columns
    wb = load_workbook(file_name)
    ws = wb.active

    # Set column formats
    for cell in ws["A"][1:]:  # Column A (Date), skip header
        cell.number_format = "DD/MM/YYYY"

    for cell in ws["H"][1:]:  # Column H (Rate), skip header
        cell.number_format = 'Currency'

    # Save formatted file
    wb.save(file_name)
    print(f"Formatted schedule saved as {file_name}")


# Example usage
generate_schedule(2024)
